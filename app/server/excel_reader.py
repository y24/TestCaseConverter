"""
Excel読み取り機能
"""
import re
import logging
import io
from pathlib import Path
from typing import List, Dict, Optional, Tuple, Union
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .models import (
    ConversionSettings, FileData, SheetData, TestCase
)

logger = logging.getLogger(__name__)


class ExcelReader:
    """Excel読み取りクラス"""
    
    def __init__(self, settings: ConversionSettings):
        self.settings = settings
        self.global_test_case_counter = (settings.id_start_number or 1) - 1  # グローバルなテストケースカウンター（開始番号-1で初期化）
    
    def read_file(self, file_source: Union[Path, bytes], filename: str) -> FileData:
        """Excelファイルを読み取り（ファイルパスまたはバイトデータから）"""
        workbook = None
        try:
            # ファイルサイズ確認
            if isinstance(file_source, Path):
                # ファイルパスから読み取り
                if not file_source.exists():
                    logger.error(f"File does not exist: {file_source}")
                    return FileData(
                        filename=filename,
                        sheets=[],
                        warnings=[f"ファイルが存在しません: {file_source}"]
                    )
                
                file_size = file_source.stat().st_size
                if file_size == 0:
                    logger.error(f"File is empty: {file_source}")
                    return FileData(
                        filename=filename,
                        sheets=[],
                        warnings=[f"ファイルが空です: {file_source}"]
                    )
                
                logger.info(f"Excel file reading started: {filename} (size: {file_size} bytes)")
                
                # ファイル形式の事前チェック
                if not self._is_valid_excel_file(file_source):
                    logger.error(f"Invalid Excel file format: {filename}")
                    return FileData(
                        filename=filename,
                        sheets=[],
                        warnings=[f"無効なExcelファイル形式です: {filename}"]
                    )
                
                workbook = load_workbook(file_source, data_only=True)
                
            else:
                # バイトデータから読み取り
                if len(file_source) == 0:
                    logger.error(f"File is empty: {filename}")
                    return FileData(
                        filename=filename,
                        sheets=[],
                        warnings=[f"ファイルが空です: {filename}"]
                    )
                
                logger.info(f"Excel file reading started: {filename} (size: {len(file_source)} bytes)")
                
                # バイトデータの形式チェック
                if not self._is_valid_excel_bytes(file_source):
                    logger.error(f"Invalid Excel file format: {filename}")
                    return FileData(
                        filename=filename,
                        sheets=[],
                        warnings=[f"無効なExcelファイル形式です: {filename}"]
                    )
                
                workbook = load_workbook(io.BytesIO(file_source), data_only=True)
            
            sheets_data = []
            file_warnings = []
            
            # 対象シートを検索
            target_sheets = self._find_target_sheets(workbook)
            
            if not target_sheets:
                file_warnings.append("対象シートが見つかりません")
                return FileData(
                    filename=filename,
                    sheets=[],
                    warnings=file_warnings
                )
            
            logger.info(f"Target sheets found: {target_sheets}")
            
            # 各シートを処理
            for sheet_name in target_sheets:
                try:
                    sheet_data = self._read_sheet(workbook[sheet_name], sheet_name)
                    sheets_data.append(sheet_data)
                except Exception as e:
                    logger.error(f"Sheet {sheet_name} reading error: {e}")
                    file_warnings.append(f"シート {sheet_name} の読み取りに失敗: {str(e)}")
            
            logger.info(f"Excel file reading completed: {filename} (sheets: {len(sheets_data)})")
            
            return FileData(
                filename=filename,
                sheets=sheets_data,
                warnings=file_warnings
            )
            
        except PermissionError as e:
            logger.error(f"File {filename} access permission error: {e}")
            return FileData(
                filename=filename,
                sheets=[],
                warnings=[f"ファイルアクセス権限エラー: {str(e)}"]
            )
        except Exception as e:
            logger.error(f"File {filename} reading error: {e}")
            # より詳細なエラー情報を提供
            error_detail = str(e)
            if "could not read worksheets from None" in error_detail:
                error_detail = "Excelファイルの構造が破損しているか、無効なXMLが含まれています。ファイルを再保存してから再度お試しください。"
            elif "Permission denied" in error_detail:
                error_detail = "ファイルが他のアプリケーションで開かれている可能性があります。ファイルを閉じてから再度お試しください。"
            elif "Invalid file format" in error_detail:
                error_detail = "ファイル形式が正しくありません。.xlsxまたは.xlsファイルをアップロードしてください。"
            
            return FileData(
                filename=filename,
                sheets=[],
                warnings=[f"ファイル読み取りエラー: {error_detail}"]
            )
        finally:
            # ワークブックを明示的に閉じる
            if workbook is not None:
                try:
                    workbook.close()
                except Exception as e:
                    logger.warning(f"Workbook close error: {e}")
    
    def _find_target_sheets(self, workbook) -> List[str]:
        """対象シートを検索"""
        target_sheets = []
        
        logger.info(f"Sheet search keys: {self.settings.sheet_search_keys}")
        logger.info(f"Sheet search ignores: {self.settings.sheet_search_ignores}")
        logger.info(f"Available sheets: {workbook.sheetnames}")
        
        # 検索キーに*が含まれている場合はすべてのシートを対象にする
        if "*" in self.settings.sheet_search_keys:
            logger.info("Found '*' in search keys, selecting all sheets")
            # *のみが設定されている場合、または*が含まれている場合はすべてのシートを対象にする
            for sheet_name in workbook.sheetnames:
                # 除外シートチェック
                if sheet_name in self.settings.sheet_search_ignores:
                    logger.info(f"Excluding sheet: {sheet_name}")
                    continue
                target_sheets.append(sheet_name)
                logger.info(f"Added sheet: {sheet_name}")
            logger.info(f"Final target sheets (with *): {target_sheets}")
            return target_sheets
        
        # 通常の検索キーによる検索
        logger.info("Using normal search keys")
        for sheet_name in workbook.sheetnames:
            # 除外シートチェック
            if sheet_name in self.settings.sheet_search_ignores:
                logger.info(f"Excluding sheet: {sheet_name}")
                continue
            
            # 検索キーチェック
            for search_key in self.settings.sheet_search_keys:
                if search_key in sheet_name:
                    target_sheets.append(sheet_name)
                    logger.info(f"Added sheet: {sheet_name} (matched key: {search_key})")
                    break
        
        logger.info(f"Final target sheets (normal): {target_sheets}")
        return target_sheets
    
    def _get_a1_cell_value(self, worksheet) -> str:
        """A1セルの値を取得"""
        try:
            a1_cell = worksheet.cell(row=1, column=1)
            if a1_cell.value is not None:
                value = str(a1_cell.value).strip()
                logger.info(f"A1セルの値を取得: '{value}'")
                return value
            logger.info("A1セルが空です")
            return ""
        except Exception as e:
            logger.warning(f"A1セルの取得に失敗: {e}")
            return ""
    
    def _read_sheet(self, worksheet, sheet_name: str) -> SheetData:
        """シートを読み取り"""
        warnings = []
        
        # A1セルの値を取得
        a1_cell_value = self._get_a1_cell_value(worksheet)
        
        # ヘッダー行を検索
        header_row = self._find_header_row(worksheet)
        if header_row is None:
            warnings.append("ヘッダー行が見つかりません")
            logger.info(f"Creating SheetData (no header) - sheet_name: '{sheet_name}', a1_cell_value: '{a1_cell_value}'")
            sheet_data = SheetData(
                sheet_name=sheet_name,
                items=[],
                warnings=warnings,
                a1_cell_value=a1_cell_value
            )
            logger.info(f"SheetData created (no header) - a1_cell_value: '{sheet_data.a1_cell_value}'")
            return sheet_data
        
        # 列マッピングを取得
        column_mapping = self._get_column_mapping(worksheet, header_row)
        if not column_mapping:
            warnings.append("必須列が見つかりません")
            logger.info(f"Creating SheetData (no mapping) - sheet_name: '{sheet_name}', a1_cell_value: '{a1_cell_value}'")
            sheet_data = SheetData(
                sheet_name=sheet_name,
                items=[],
                warnings=warnings,
                a1_cell_value=a1_cell_value
            )
            logger.info(f"SheetData created (no mapping) - a1_cell_value: '{sheet_data.a1_cell_value}'")
            return sheet_data
        
        # 新しいセル情報を取得
        cell_info = self._get_cell_info(worksheet, header_row)
        
        # テスト環境情報を取得
        test_environments = self._get_test_environments(worksheet)
        
        # データ行を取得
        data_rows = self._get_data_rows(worksheet, header_row, column_mapping)
        
        # テストケースに変換
        test_cases = []
        for row_data in data_rows:
            try:
                test_case = self._create_test_case(row_data, column_mapping, sheet_name, cell_info, test_environments)
                if test_case:
                    test_cases.append(test_case)
            except Exception as e:
                logger.error(f"Test case creation error (row {row_data.get('row', '?')}): {e}")
                warnings.append(f"行 {row_data.get('row', '?')} の処理に失敗: {str(e)}")
        
        logger.info(f"Creating SheetData - sheet_name: '{sheet_name}', a1_cell_value: '{a1_cell_value}'")
        sheet_data = SheetData(
            sheet_name=sheet_name,
            items=test_cases,
            warnings=warnings,
            a1_cell_value=a1_cell_value
        )
        logger.info(f"SheetData created - a1_cell_value: '{sheet_data.a1_cell_value}'")
        return sheet_data
    
    def _find_header_row(self, worksheet) -> Optional[int]:
        """ヘッダー行を検索"""
        search_col = self.settings.header.search_col
        search_key = self.settings.header.search_key
        
        # 列番号に変換
        if search_col.isalpha():
            col_num = ord(search_col.upper()) - ord('A') + 1
        else:
            col_num = int(search_col)
        
        # 最大50行まで検索
        for row in range(1, min(51, worksheet.max_row + 1)):
            cell_value = worksheet.cell(row=row, column=col_num).value
            if cell_value and search_key in str(cell_value):
                return row
        
        return None
    
    def _get_column_mapping(self, worksheet, header_row: int) -> Dict[str, int]:
        """列マッピングを取得"""
        mapping = {}
        
        # ヘッダー行の各列をチェック
        for col in range(1, worksheet.max_column + 1):
            header_value = worksheet.cell(row=header_row, column=col).value
            if not header_value:
                continue
            
            header_str = str(header_value).strip()
            
            # カテゴリ列の個別マッピング
            for i, category_key in enumerate(self.settings.category_row.keys):
                if category_key in header_str:
                    mapping[f'category_{i}'] = col
                    break
            
            # その他の設定のキーと照合
            configs = [
                ('step', self.settings.step_row),
                ('tobe', self.settings.tobe_row),
                ('test_type', self.settings.test_type_row),
                ('priority', self.settings.priority_row),
                ('precondition', self.settings.precondition_row),
                ('note', self.settings.note_row),
            ]
            
            for config_name, config in configs:
                for key in config.keys:
                    if key in header_str:
                        # 除外チェック
                        should_exclude = False
                        for ignore_key in config.ignores:
                            if ignore_key in header_str:
                                should_exclude = True
                                break
                        
                        if not should_exclude:
                            mapping[config_name] = col
                            break
        
        return mapping
    
    def _get_cell_info(self, worksheet, header_row: int) -> Dict[str, str]:
        """検索列から新しいセル情報を取得"""
        cell_info = {}
        search_col = self.settings.header.search_col
        
        # 列番号に変換
        if search_col.isalpha():
            col_num = ord(search_col.upper()) - ord('A') + 1
        else:
            col_num = int(search_col)
        
        logger.info(f"Searching for cell info in column {search_col} (col_num: {col_num})")
        
        # 検索対象の設定
        cell_configs = [
            ('backlog_id', self.settings.backlog_id_cell),
            ('test_type', self.settings.test_type_cell),
            ('test_target', self.settings.test_target_cell),
            ('target_version', self.settings.target_version_cell),
        ]
        
        logger.info(f"Cell configs: {[(name, config.keys) for name, config in cell_configs]}")
        
        # 最大50行まで検索
        for row in range(1, min(51, worksheet.max_row + 1)):
            cell_value = worksheet.cell(row=row, column=col_num).value
            if not cell_value:
                continue
                
            cell_str = str(cell_value).strip()
            logger.debug(f"Row {row}, Column {col_num}: '{cell_str}'")
            
            # 各設定のキーと照合
            for info_name, config in cell_configs:
                if info_name in cell_info:  # 既に見つかっている場合はスキップ
                    continue
                    
                for key in config.keys:
                    if key in cell_str:
                        logger.info(f"Found key '{key}' in cell '{cell_str}' at row {row}")
                        # 除外チェック
                        should_exclude = False
                        for ignore_key in config.ignores:
                            if ignore_key in cell_str:
                                should_exclude = True
                                break
                        
                        if not should_exclude:
                            # 2つ右隣のセルを取得
                            target_col = col_num + 2
                            target_cell_value = worksheet.cell(row=row, column=target_col).value
                            if target_cell_value is not None:
                                cell_info[info_name] = str(target_cell_value).strip()
                                logger.info(f"Found {info_name}: '{cell_info[info_name]}' at row {row}, column {target_col}")
                            else:
                                cell_info[info_name] = ""
                                logger.info(f"Found {info_name} but target cell is empty at row {row}, column {target_col}")
                            break
        
        logger.info(f"Final cell_info: {cell_info}")
        return cell_info
    
    def _get_test_environments(self, worksheet) -> List[str]:
        """1行目からテスト環境情報を取得（A列を除外、除外設定を適用）"""
        test_environments = []
        
        # 1行目を検索（A列は除外）
        for col in range(2, worksheet.max_column + 1):  # B列から開始
            cell_value = worksheet.cell(row=1, column=col).value
            if cell_value is not None:
                cell_str = str(cell_value).strip()
                if cell_str:  # 空でない文字列の場合
                    # セル内改行を半角スペースに置換
                    import re
                    cell_str = re.sub(r'\r\n|\r|\n', ' ', cell_str)
                    
                    # 除外チェック
                    should_exclude = False
                    for ignore_key in self.settings.test_environments_cell.ignores:
                        if ignore_key in cell_str:
                            should_exclude = True
                            logger.info(f"Excluding test environment '{cell_str}' due to ignore key '{ignore_key}'")
                            break
                    
                    if not should_exclude:
                        test_environments.append(cell_str)
                        logger.info(f"Found test environment: '{cell_str}' at row 1, column {col}")
        
        logger.info(f"Found test environments: {test_environments}")
        return test_environments
    
    def _get_data_rows(self, worksheet, header_row: int, column_mapping: Dict[str, int]) -> List[Dict]:
        """データ行を取得"""
        data_rows = []
        
        # 期待結果列の最後の行を検索
        end_row = self._find_end_row(worksheet, column_mapping.get('tobe', 1))
        
        # カテゴリのforward fill用のバッファ
        last_category_values = [""] * len(self.settings.category_row.keys)
        
        # データ行を取得
        for row in range(header_row + 1, end_row + 1):
            row_data = {'row': row}
            
            # 各列の値を取得
            for col_name, col_num in column_mapping.items():
                cell_value = worksheet.cell(row=row, column=col_num).value
                if cell_value is not None:
                    # セル値を文字列に変換し、前後の空白を削除
                    cell_str = str(cell_value)
                    # 各種改行文字を統一
                    import re
                    cell_str = re.sub(r'\r\n|\r', '\n', cell_str)
                    row_data[col_name] = cell_str.strip()
                else:
                    row_data[col_name] = ""
            
            # カテゴリを統合
            category_values = []
            for i in range(len(self.settings.category_row.keys)):
                category_key = f'category_{i}'
                if category_key in row_data:
                    category_values.append(row_data[category_key])
                else:
                    category_values.append("")
            
            # Forward fill処理
            if self.settings.forward_fill_category:
                category_values = self._forward_fill_category(category_values, last_category_values)
            
            row_data['category'] = category_values
            
            # 期待結果が空でない行のみ追加
            if row_data.get('tobe', '').strip():
                data_rows.append(row_data)
                # 有効な行の場合のみ、カテゴリバッファを更新
                last_category_values = category_values.copy()
        
        return data_rows
    
    def _find_end_row(self, worksheet, tobe_col: int) -> int:
        """期待結果列の最後の行を検索"""
        # ヘッダー行を取得
        header_row = self._find_header_row(worksheet)
        if header_row is None:
            return 1
        
        end_row = header_row + 1
        
        # 下から上に検索
        for row in range(worksheet.max_row, header_row, -1):
            cell_value = worksheet.cell(row=row, column=tobe_col).value
            if cell_value and str(cell_value).strip():
                end_row = row
                break
        
        return end_row
    
    def _forward_fill_category(self, current_category: List[str], last_category: List[str]) -> List[str]:
        """カテゴリの空の値を前の行から埋める（独立ブランチ対応）"""
        filled_category = []
        
        # 左側のカテゴリが変わったかどうかをチェック
        independent_branch = False
        for i, current_value in enumerate(current_category):
            # ダッシュ文字（- や －）を無視する
            if self._is_dash_character(current_value):
                # ダッシュ文字の場合は前の行の値を使用（独立ブランチでない場合）
                if independent_branch:
                    filled_category.append("")
                else:
                    if i < len(last_category):
                        filled_category.append(last_category[i])
                    else:
                        filled_category.append("")
            elif current_value.strip():  # 現在の値が空でない場合
                filled_category.append(current_value)
                # このレベルで値が変わった場合、それより右側は独立ブランチ
                if i < len(last_category) and current_value != last_category[i]:
                    independent_branch = True
            else:  # 現在の値が空の場合
                if independent_branch:
                    # 独立ブランチの場合は空のまま
                    filled_category.append("")
                else:
                    # 通常の場合は前の行の値を使用
                    if i < len(last_category):
                        filled_category.append(last_category[i])
                    else:
                        filled_category.append("")
        
        # 独立ブランチが発生した場合、残りの要素も空にする
        while len(filled_category) < len(current_category):
            filled_category.append("")
        
        return filled_category
    
    def _is_dash_character(self, value: str) -> bool:
        """ダッシュ文字（- や －）かどうかを判定"""
        if not value:
            return False
        
        # 前後の空白を削除
        trimmed_value = value.strip()
        
        # ダッシュ文字のパターンをチェック
        dash_patterns = ['-', '－', 'ー', '―', '—', '–']
        
        return trimmed_value in dash_patterns
    
    def _create_test_case(self, row_data: Dict, column_mapping: Dict[str, int], sheet_name: str, cell_info: Dict[str, str] = None, test_environments: List[str] = None) -> Optional[TestCase]:
        """テストケースを作成"""
        # カテゴリを取得（既にリスト形式で統合済み）
        category = row_data.get('category', [])
        
        # 手順を文字列として取得（前後の空白を削除、連続する改行を単一の改行に正規化）
        steps = self._normalize_multiline_text(row_data.get('step', ''))
        
        # 期待結果を文字列として取得（前後の空白を削除、連続する改行を単一の改行に正規化）
        expect = self._normalize_multiline_text(row_data.get('tobe', ''))
        
        # 前提条件を文字列として取得（前後の空白を削除、連続する改行を単一の改行に正規化）
        preconditions = self._normalize_multiline_text(row_data.get('precondition', ''))
        
        # タイトルを文字列として取得（前後の空白を削除、連続する改行を単一の改行に正規化）
        title = self._normalize_multiline_text(row_data.get('title', ''))
        
        # テストケースIDを生成（グローバルカウンターを使用して通しの連番）
        self.global_test_case_counter += 1
        test_id = f"{self.settings.id_prefix}-{self.global_test_case_counter:0{self.settings.id_padding}d}"
        
        # 新しいフィールドの値を取得
        backlog_id = cell_info.get('backlog_id', '') if cell_info else ''
        test_type = cell_info.get('test_type', '') if cell_info else ''
        test_target = cell_info.get('test_target', '') if cell_info else ''
        target_version = cell_info.get('target_version', '') if cell_info else ''
        
        # テスト環境情報を取得
        test_envs = test_environments if test_environments else []
        
        return TestCase(
            id=test_id,
            title=title,
            category=category,
            type=row_data.get('test_type', ''),
            priority=row_data.get('priority', ''),
            preconditions=preconditions,
            steps=steps,
            expect=expect,
            notes=row_data.get('note', ''),
            source={
                'file': sheet_name,
                'sheet': sheet_name,
                'row': row_data['row']
            },
            # 新しいフィールド
            backlog_id=backlog_id,
            test_type=test_type,
            test_target=test_target,
            target_version=target_version,
            test_environments=test_envs
        )
    
    
    
    
    def _normalize_multiline_text(self, text: str) -> str:
        """複数行テキストを正規化（余計な空白行を削除）"""
        if not text:
            return ""
        
        # 前後の空白を削除
        text = text.strip()
        
        # 各種改行文字を統一（\r\n, \r, \n を \n に統一）
        import re
        text = re.sub(r'\r\n|\r', '\n', text)
        
        # 改行で分割
        lines = text.split('\n')
        
        # 各行の前後の空白を削除し、空行を除去
        normalized_lines = []
        for line in lines:
            line = line.strip()
            if line:  # 空行でない場合のみ追加
                normalized_lines.append(line)
        
        # 正規化された行を結合
        result = '\n'.join(normalized_lines)
        
        return result
    
    def _is_valid_excel_file(self, file_path: Path) -> bool:
        """Excelファイルの形式を検証（ファイルパス版）"""
        try:
            # ファイルの先頭バイトをチェック
            with open(file_path, 'rb') as f:
                header = f.read(8)
                # ZIPファイルのマジックナンバー（ExcelファイルはZIPベース）
                if header[:4] == b'PK\x03\x04':
                    return True
                # 古いExcel形式のマジックナンバー
                elif header[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
                    return True
                return False
        except Exception:
            return False
    
    def _is_valid_excel_bytes(self, file_bytes: bytes) -> bool:
        """Excelファイルの形式を検証（バイトデータ版）"""
        try:
            if len(file_bytes) < 8:
                return False
            
            # ZIPファイルのマジックナンバー（ExcelファイルはZIPベース）
            if file_bytes[:4] == b'PK\x03\x04':
                return True
            # 古いExcel形式のマジックナンバー
            elif file_bytes[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
                return True
            return False
        except Exception:
            return False
    
    
